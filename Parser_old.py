# Импорт библиотек
import os.path
import pandas as pd
import openpyxl
from openpyxl import load_workbook

start_year = '1990'
end_year = '2019'
file_to_parse = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'

def parser_flow():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)
    resources = ['Electricity', 'Natural gas', 'Coal, peat and oil shale', 'Oil products', 'Heat']
    flow = 'Industry (PJ)'
    file = r'C:\Users\Артем\Desktop\МФТИ\Магистратура\Диплом_Магистр\World Energy Balances Highlights 2021.xlsx'

    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле

    # Загружаем лист (sheet_name) в DataFrame 'df', пропуская (по желанию) строки (skiprows) или столбцы (skipcols)
    df = xl.parse(sheet_name='TimeSeries_1971-2020', skiprows=1)

    # print(df.keys()) # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме

    if os.path.exists(file_to_parse):  # Проверяем, есть ли файл, чтобы задать нужные параметры для записи
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    # mode = "a" if os.path.exists(file_to_parse) else "w"
    writer = pd.ExcelWriter(file_to_parse, mode=mode, engine='openpyxl',
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки
    # Запись данных
    for i in range(0, 5, 1):
        df1 = df[(df['Product'] == resources[i]) & (df['Flow'] == flow)]
        # Фильтрация данных
        country = (df1['Country'])  # Отбор нужных столбцов
        for k in range(int(start_year), int(end_year) + 1, 1):
            data = (df1[k])  # Отбор нужных столбцов
            data.name = resources[i]  # Переименовываем столбец (rename не работает, т.к. здесь он всего один)
            # Запись в новый Excel-файл
            # book = load_workbook(file_to_parse)  # Получаем доступ к файлу MS Excel в который будем записывать датафрейм
            # writer.book = book  # Сохраняем предыдущую информацию файла, чтобы при записи она осталась
            # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)  # ExcelWriter использует эту переменную для доступа к листу.
            # Если оставить ее пустой, он не будет знать, что лист уже существует, и создаст новый лист.

            country.to_excel(writer, sheet_name=str(k), index=False, startcol=0)  # Записываем датафрейм в файл.
            # sheet_name='2019' показывает, в какой лист записываем.
            data.to_excel(writer, sheet_name=str(k), index=False, startcol=i + 1)  # Записываем датафрейм в файл.
            # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).

    writer.save()  # Сохраняем результат

    # # Удаление страницы по умолчанию (уже не нужно, так как writer создает пустой файл Excel)
    # wb = load_workbook(file_to_parse)  # Получаем доступ к файлу MS Excel в который записали датафрейм
    # sheet = wb.sheetnames  # Получили список всех листов в файле и загнали его в переменную
    # #print(sheet)  # Вывели на экран список всех листов в файле
    # pfd = wb['Лист1']  # Сделали активной страницу, которую хотим удалить, где ['Лист1'] - название страницы.
    # wb.remove(pfd)  # Удаляем эту страницу
    # wb.save(file_to_parse)  # Сохранили файл с изменениями (удаленная страница)

parser_flow()

def value_added():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)

    file = r'C:\Users\Артем\Desktop\МФТИ\Магистратура\Диплом_Магистр\Industry (include construction) value added WB.xlsx'

    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле

    # Загружаем лист (sheet_name) в DataFrame 'df', пропуская (по желанию) строки (skiprows) или столбцы (skipcols)
    df = xl.parse(sheet_name='Data', skiprows=3)
    # print(df.keys()) # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме

    # Запись данных
    # Фильтрация данных
    country = (df['Country Name'])  # Отбор нужных столбцов

    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, mode=mode, engine='openpyxl',
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки

    for k in range(int(start_year), int(end_year) + 1, 1):
        data = (df[str(k)])  # Отбор нужных столбцов
        data.name = 'Value added'  # Переименовываем столбец (rename не работает, т.к. здесь он всего один)
        # # Запись в новый Excel-файл
        # book = load_workbook(file_to_parse)  # Получаем доступ к файлу MS Excel в который будем записывать датафрейм
        # writer.book = book  # Сохраняем предыдущую информацию файла, чтобы при записи она осталась
        # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)  # ExcelWriter использует эту переменную для доступа к листу.
        # # Если оставить ее пустой, он не будет знать, что лист уже существует, и создаст новый лист.

        country.to_excel(writer, sheet_name=str(k), index=False, startcol=10)  # Записываем датафрейм в файл.
        # sheet_name='2019' показывает, в какой лист записываем.
        data.to_excel(writer, sheet_name=str(k), index=False, startcol=11)  # Записываем датафрейм в файл.
        # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

value_added()

def useful_cons():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)
    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл
    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        df['Useful consumption'] = ((df['Electricity'] + df['Natural gas'] + df['Coal, peat and oil shale']) * 0.35 + (
                df['Oil products'] + df['Heat']) * 0.9)
        df1 = df['Useful consumption']
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        df1.to_excel(writer, sheet_name=str(k), index=False, startcol=6)  # Записываем датафрейм в файл.
    # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

useful_cons()