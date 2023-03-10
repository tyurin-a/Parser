# Импорт библиотек
import os.path
import pandas as pd
from tqdm import trange
import numpy as np
from numpy import nan
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, LineChart, Series
from openpyxl.chart.label import DataLabelList

start_year = '1990'
end_year = '2019'
file_to_parse = r'C:\Users\Артем\Desktop\Total energy supply.xlsx'
resources = ['Coal and coal products', 'Peat and peat products', 'Oil shale and oil sands', 'Crude, NGL and feedstocks',
             'Oil products', 'Natural gas', 'Nuclear', 'Hydro', 'Geothermal', 'Solar/wind/other', 'Biofuels and waste',
             'Heat production from non-specified combustible fuels', 'Electricity',  'Heat', 'Total']



def total2():
    file = r'C:\Users\Артем\Desktop\Total energy supply.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл

    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки

    dfc = dfs[start_year]
    country = dfc['COUNTRY']
    country.to_excel(writer, sheet_name='Total2', index=False, startcol=0)
    i = 1
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        df1 = df['Energy Intensity2']
        df1.name = k
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        df1.to_excel(writer, sheet_name='Total2', index=False, startcol=i)  # Записываем датафрейм в файл.
        # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
        i += 1
    writer.save()  # Сохраняем результат

total2()
print('Общий лист 2 сформирован')

def normalize2():
    file = r'C:\Users\Артем\Desktop\Total energy supply.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    df = []
    df = xl.parse(sheet_name='Total2', skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл

    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки

    df.reset_index(drop=True)  # Cбрасываем индексирование строк
    df['Max'] = df.max(axis=1,
                       numeric_only=True)  # Находим максимальное значение в каждой строке среди численных данных
    df1 = df['Max']
    country = df['COUNTRY']
    df2 = df.loc[:, start_year:end_year]
    df1.to_excel(writer, sheet_name='Total2', index=False,
                 startcol=31)  # Записываем столбец с максимальным значением в файл.
    country.to_excel(writer, sheet_name='Normal2', index=False,
                     startcol=0)  # Записываем столбец стран для нормированной таблицы.

    # Нормируем таблицу на максимальное значение
    for i in trange(0, 160):
        val = df2.iloc[i]  # Выбираем строку значений из df2
        max_val = df1.iloc[i]  # Выбираем строку значений из df1 (максимальное значение)
        df2.iloc[i] = val / max_val
        df2.to_excel(writer, sheet_name='Normal2', index=False, startcol=1)

    writer.save()  # Сохраняем результат

normalize2()
print('Значения в общем листе 2 отнормированы')

def plot2():
    file = r'C:\Users\Артем\Desktop\Total energy supply.xlsx'

    df = openpyxl.load_workbook(file)  # Читаем файл
    sheet = df['Normal2']  # Выбираем нужный лист

    chart = LineChart()  # Создаем объект LineChart

    # countries = Reference(sheet, min_col=34, max_col=34, min_row=2, max_row=159)
    years = Reference(sheet, min_col=2, max_col=31, min_row=1,
                      max_row=1)  # Подаем список годов, по которому будет определяться ось х на графике
    # data = Reference(sheet, min_col=35, max_col=64, min_row=2, max_row=159)
    # Записываем легенду графика, а также определяем данные, по которым строится сам график
    for i in range(2, 160):
        chart.series.append(
            Series(Reference(sheet, min_col=2, max_col=31, min_row=i, max_row=i), title=sheet.cell(i, 1).value))
    # chart.add_data(data, from_rows=True)
    chart.set_categories(years)  # Указываем, какой должна быть ось х на графике
    chart.width = 30  # Ширина и высота графика (в см)
    chart.height = 10

    sheet.add_chart(chart, "E5")  # Добавляем график на лист в переменной sheet с левым верхним углом в ячейке Е5
    df.save(file)

plot2()
print('Общий график 2 (без разбиения стран на группы) построен')
