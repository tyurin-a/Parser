# Импорт библиотек
import os.path
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook

start_year = '1990'
end_year = '2019'
file_to_parse = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
resources = ['Coal and coal products', 'Oil products', 'Natural gas', 'Electricity',  'Heat']

def parser_flow():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)
    file = r'C:\Users\Артем\Desktop\Tw.xlsx'

    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле

    # Загружаем лист (sheet_name) в DataFrame 'df', пропуская (по желанию) строки (skiprows) или столбцы (skipcols)
    df = xl.parse(sheet_name='Tw', skiprows=3)

    # print(df.keys()) # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме

    if os.path.exists(file_to_parse):  # Проверяем, есть ли файл, чтобы задать нужные параметры для записи
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    # mode = "a" if os.path.exists(file_to_parse) else "w"
    writer = pd.ExcelWriter(file_to_parse, mode=mode, engine='openpyxl',
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки
    # Запись данных
    country = df.loc[df['PRODUCT'] == resources[0], 'COUNTRY']  # Отбор нужных столбцов
    for i in range(0, 5, 1):
        df1 = df[(df['PRODUCT'] == resources[i])]
        # Фильтрация данных
        for k in range(int(start_year), int(end_year) + 1, 1):
            data = (df1[str(k)])  # Отбор нужных столбцов
            data.name = resources[i]  # Переименовываем столбец (rename не работает, т.к. здесь он всего один)
            # Запись в новый Excel-файл
            # book = load_workbook(file_to_parse)  # Получаем доступ к файлу MS Excel в который будем записывать датафрейм
            # writer.book = book  # Сохраняем предыдущую информацию файла, чтобы при записи она осталась
            # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)  # ExcelWriter использует эту переменную для доступа к листу.
            # Если оставить ее пустой, он не будет знать, что лист уже существует, и создаст новый лист.
            #df1[str(k)] = np.where(((df1[str(k)] == 'x') | (df1[str(k)] == '..')), 0, df1[str(k)])
            country.to_excel(writer, sheet_name=str(k), index=False, startcol=0)  # Записываем датафрейм в файл.
            # sheet_name='2019' показывает, в какой лист записываем.
            data.to_excel(writer, sheet_name=str(k), index=False, startcol=i + 1)  # Записываем датафрейм в файл.
            # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).

    writer.save()  # Сохраняем результат

    # # Удаление страницы по умолчанию (уже не нужно, так как writer создает пустой файл Excel)
    # wb = load_workbook(file_to_parse)  # Получаем доступ к файлу MS Excel в который записали датафрейм
    # sheet = wb.sheetnames  # Получили список всех листов в файле и загнали его в переменную
    # #print(sheet)  # Вывели на экран список всех листов в файле
    # pfd = wb['Лист1']  # Сделали активной страницу, которую хотим удалить, где ['Лист1'] - название страницы.
    # wb.remove(pfd)  # Удаляем эту страницу
    # wb.save(file_to_parse)  # Сохранили файл с изменениями (удаленная страница)

parser_flow()

def remove_symb():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)
    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл
    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        for i in range(0, 5, 1):
            df.loc[df[resources[i]] == 'x', resources[i]] = 0
            df.loc[df[resources[i]] == '..', resources[i]] = 0
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        df.to_excel(writer, sheet_name=str(k), index=False, startcol=0)  # Записываем датафрейм в файл.
    # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

remove_symb()

def parser_va():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)

    file = r'C:\Users\Артем\Desktop\МФТИ\Магистратура\Диплом_Магистр\Industry (include construction) value added WB.xlsx'

    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле

    # Загружаем лист (sheet_name) в DataFrame 'df', пропуская (по желанию) строки (skiprows) или столбцы (skipcols)
    df = xl.parse(sheet_name='Data', skiprows=3)
    # print(df.keys()) # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме

    # Запись данных
    # Фильтрация данных
    country = (df['Country Name'])  # Отбор нужных столбцов

    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, mode=mode, engine='openpyxl',
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки

    for k in range(int(start_year), int(end_year) + 1, 1):
        data = (df[str(k)])  # Отбор нужных столбцов
        data.name = 'Value added'  # Переименовываем столбец (rename не работает, т.к. здесь он всего один)
        # # Запись в новый Excel-файл
        # book = load_workbook(file_to_parse)  # Получаем доступ к файлу MS Excel в который будем записывать датафрейм
        # writer.book = book  # Сохраняем предыдущую информацию файла, чтобы при записи она осталась
        # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)  # ExcelWriter использует эту переменную для доступа к листу.
        # # Если оставить ее пустой, он не будет знать, что лист уже существует, и создаст новый лист.

        country.to_excel(writer, sheet_name=str(k), index=False, startcol=10)  # Записываем датафрейм в файл.
        # sheet_name='2019' показывает, в какой лист записываем.
        data.to_excel(writer, sheet_name=str(k), index=False, startcol=11)  # Записываем датафрейм в файл.
        # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

parser_va()

def useful_cons():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)
    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл
    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        df['Useful consumption'] = ((df['Electricity'] + df['Natural gas'] + df['Coal and coal products']) * 0.35 +
                                    (df['Oil products'] + df['Heat']) * 0.9)
        df1 = df['Useful consumption']
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        df1.to_excel(writer, sheet_name=str(k), index=False, startcol=6)  # Записываем датафрейм в файл.
    # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

useful_cons()

def change_keys():
    # Импорт библиотек
    import os.path
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook

    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл

    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл
    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки
    country_dict = {"China": "People's Republic of China", "Bolivia": "Plurinational State of Bolivia", "Congo, Rep.": "Republic of the Congo", "Congo, Dem. Rep.": "Democratic Republic of the Congo", "Cote d'Ivoire" : "Cфte d'Ivoire", "Curacao": "Curaзao/Netherlands Antilles", "Korea, Dem. People's Rep.": "Democratic People's Republic of Korea",
                    "Egypt, Arab Rep.": "Egypt", "Hong Kong SAR, China": "Hong Kong (China)", "Iran, Islamic Rep.": "Islamic Republic of Iran", "Korea, Rep.": "Korea", "Kyrgyz Republic": "Kyrgyzstan", "Lao PDR": "Lao People's Democratic Republic", "Moldova": "Republic of Moldova", "North Macedonia": "Republic of North Macedonia", "Tanzania": "United Republic of Tanzania",
                    "Turkiye": "Turkey", "Venezuela, RB": "Bolivarian Republic of Venezuela", "Vietnam": "Viet Nam", "Yemen, Rep.": "Yemen"
                    }  # Словарь сравнительной таблицы

    cn = pd.DataFrame(list(country_dict.items()), columns=['TableWB', 'TableIEA'])  # Создаем датафрейм из словаря
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        table2 = df.loc[:, 'Country Name':'Value added']
        table2 = pd.merge(table2, cn, left_on=['Country Name'], right_on=['TableWB'],
                          how='left')  # Аналог ВПР для таблицы 2,
        # в которой будем менять ключи для адекватного соединения с 1ой таблицей с помощью сравнительной таблицы cn
        table2.drop(['TableWB'], axis='columns', inplace=True)
        table2.loc[pd.notna(table2['TableIEA']) == True, 'Country Name'] = table2['TableIEA']  # Меняем ключи в столбце 'Country Name'
        # таблицы 2 на значения столбца 'TableIEA', там где значение в строках столбца 'TableIEA' не пусто
        table2.drop(['TableIEA'], axis='columns', inplace=True)  # Удаляем лишний столбец
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        table2.to_excel(writer, sheet_name=str(k), index=False, startcol=10)
    # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

change_keys()

def vlookup():
    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл
    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки

    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        table1 = df.loc[:, 'COUNTRY':'Useful consumption']  # Создаем датафреймы из двух таблиц на листе
        table2 = df.loc[:, 'Country Name':'Value added']
        table1 = pd.merge(table1, table2, left_on=['COUNTRY'], right_on=['Country Name'], how='left')  # Аналог ВПР
        table1.drop(['Country Name'], axis='columns', inplace=True)  # Удаляем лишний столбец
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        table1.to_excel(writer, sheet_name=str(k), index=False, startcol=0)  # Записываем датафрейм в файл.
    # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

vlookup()

def specific_en_cons():
    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode="a",
                            if_sheet_exists="overlay")  # Указываем writer библиотеки
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        table1 = df.loc[:, 'COUNTRY':'Value added']  # Создаем датафрейм
        table1['Specific energy consumption'] = (table1['Useful consumption'] / table1['Value added'])
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        table1.to_excel(writer, sheet_name=str(k), index=False, startcol=0)  # Записываем датафрейм в файл.
    # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
    writer.save()  # Сохраняем результат

specific_en_cons()

def total():
    # Задаем стартовые параметры для парсинга и фильтрации (начальный и конечный года, энергоресурсы, отрасль и путь к файлу)
    file = r'C:\Users\Артем\Desktop\Industry consumption.xlsx'
    xl = pd.ExcelFile(file)  # Загружаем spreadsheet (электронную таблицу) в объект pandas
    # print(xl.sheet_names) # Печатаем названия листов в данном файле
    dfs = {

    }  # Словарь, в который выгружаем эксель-файл
    for k in xl.sheet_names:
        dfs[k] = xl.parse(sheet_name=str(k), skiprows=0)  # Парсим листы эксель-файла
    xl.close()  # Закрываем читаемый файл

    if os.path.exists(file_to_parse):
        mode = "a"
        if_sheet_exists = "overlay"
    else:
        mode = "w"
        if_sheet_exists = None
    writer = pd.ExcelWriter(file_to_parse, engine='openpyxl', mode=mode,
                            if_sheet_exists=if_sheet_exists)  # Указываем writer библиотеки

    dfc = dfs[start_year]
    country = dfc['COUNTRY']
    country.to_excel(writer, sheet_name='Total', index=False, startcol=0)
    i = 1
    for k in dfs:
        df = dfs[k]  # Получаем лист из словаря dfs
        df1 = df['Specific energy consumption']
        df1.name = k
        # print(df1.name)  # Печатаем названия первичных ключей (названия столбцов) в данном массиве (не в датафрейме)
        # print(df.keys())  # Печатаем названия первичных ключей (названия столбцов) в данном датафрейме
        df1.to_excel(writer, sheet_name='Total', index=False, startcol=i)  # Записываем датафрейм в файл.
        # index=False отключает запись индексов, startcol=1 начианет запись с 1 стобца (нумерация с нуля).
        i += 1
    writer.save()  # Сохраняем результат

total()





